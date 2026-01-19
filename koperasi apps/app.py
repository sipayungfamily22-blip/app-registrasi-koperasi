from dotenv import load_dotenv
load_dotenv()
from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, send_file
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
import os
from werkzeug.utils import secure_filename
from datetime import datetime
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import csv
import tempfile
from io import StringIO, BytesIO
from openpyxl import load_workbook
import logging

from config import Config, DevelopmentConfig, ProductionConfig
from database import db
from models import User, Pendaftaran, Dokter, PenggunaanKupon

# Set up template and static folder paths
template_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'templates'))
static_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'static'))

app = Flask(__name__, template_folder=template_dir, static_folder=static_dir)

# Choose configuration based on environment
if os.getenv('FLASK_ENV') == 'production':
    app.config.from_object(ProductionConfig)
else:
    app.config.from_object(DevelopmentConfig)

# Setup logging
log_file = os.path.join(os.path.dirname(__file__), '..', app.config.get('LOG_FILE', 'logs/app.log'))
os.makedirs(os.path.dirname(log_file), exist_ok=True)
handler = logging.FileHandler(log_file)
handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)
app.logger.addHandler(handler)
app.logger.setLevel(logging.INFO)

# Initialize extensions
db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@app.before_request
def log_request_info():
    user_email = current_user.email if current_user.is_authenticated else 'Anonymous'
    app.logger.info(f'API Access: {request.method} {request.path} by {user_email} from {request.remote_addr}')

@app.after_request
def log_response_info(response):
    user_email = current_user.email if current_user.is_authenticated else 'Anonymous'
    app.logger.info(f'API Response: {request.method} {request.path} - Status: {response.status_code} by {user_email}')
    return response

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def save_upload(file, folder, filename_prefix):
    if file and allowed_file(file.filename):
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = secure_filename(f"{filename_prefix}_{timestamp}_{file.filename}")
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], folder, filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        file.save(filepath)
        return filename
    return None

def send_email(to_email, subject, body):
    try:
        msg = MIMEMultipart()
        msg['From'] = app.config['MAIL_DEFAULT_SENDER']
        msg['To'] = to_email
        msg['Subject'] = subject
        
        msg.attach(MIMEText(body, 'html'))
        
        server = smtplib.SMTP(app.config['MAIL_SERVER'], app.config['MAIL_PORT'])
        server.starttls()
        server.login(app.config['MAIL_USERNAME'], app.config['MAIL_PASSWORD'])
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print(f"Error sending email: {e}")
        return False

@app.route('/')
def index():
    # Get statistics
    total_pendaftar = Pendaftaran.query.count()
    approved = Pendaftaran.query.filter_by(status='approved2').count()
    pending = Pendaftaran.query.filter(Pendaftaran.status.in_(['pending', 'approved1'])).count()
    rejected = Pendaftaran.query.filter_by(status='rejected').count()
    
    stats = {
        'total_pendaftar': total_pendaftar,
        'approved': approved,
        'pending': pending,
        'rejected': rejected
    }
    
    return render_template('index.html', stats=stats)

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        app.logger.info(f'Registration attempt from {request.form.get("alamat_email", "unknown")}')
        try:
            # Validasi checkbox persyaratan
            required_checkboxes = [
                'syarat_karyawan_aktif',
                'syarat_potongan_gaji',
                'syarat_adart',
                'syarat_partisipasi',
                'pernyataan'
            ]
            
            missing_checkboxes = [cb for cb in required_checkboxes if cb not in request.form]
            if missing_checkboxes:
                flash('Anda harus menyetujui semua persyaratan dan pernyataan sebelum melanjutkan pendaftaran', 'danger')
                return redirect(url_for('register'))
            
            # Get form data
            nomor_ktp = request.form['nomor_ktp'].strip()
            
            # Validasi KTP tidak boleh terdaftar sebelumnya
            existing_ktp = Pendaftaran.query.filter_by(nomor_ktp=nomor_ktp).first()
            if existing_ktp:
                flash('Nomor KTP ini sudah pernah terdaftar. Jika merasa ada kesalahan, hubungi administrator.', 'danger')
                return redirect(url_for('register'))
            
            data = {
                'nama_lengkap': request.form['nama_lengkap'],
                'tempat_lahir': request.form['tempat_lahir'],
                'tanggal_lahir': datetime.strptime(request.form['tanggal_lahir'], '%Y-%m-%d').date(),
                'alamat_rumah': request.form['alamat_rumah'],
                'nomor_handphone': request.form['nomor_handphone'],
                'alamat_email': request.form['alamat_email'],
                'nomor_ktp': nomor_ktp,
                'no_id_karyawan': request.form['no_id_karyawan'],
                'asal_departemen': request.form['asal_departemen'],
                'nomor_rekening': request.form['nomor_rekening'],
                'nama_bank': request.form['nama_bank'],
                'pernyataan_disetujui': 'pernyataan' in request.form
            }
            
            # Save uploaded files
            data['upload_ktp'] = save_upload(request.files['upload_ktp'], 'ktp', 'ktp')
            data['upload_id_karyawan'] = save_upload(request.files['upload_id_karyawan'], 'id_karyawan', 'id_karyawan')
            data['upload_pas_foto'] = save_upload(request.files['upload_pas_foto'], 'pas_foto', 'pas_foto')
            data['upload_buku_tabungan'] = save_upload(request.files['upload_buku_tabungan'], 'buku_tabungan', 'buku_tabungan')
            
            # Create new registration
            pendaftaran = Pendaftaran(**data)
            db.session.add(pendaftaran)
            db.session.commit()
            app.logger.info(f'Registration successful for {data["nama_lengkap"]} - ID: {pendaftaran.id}')
            
            # Send notification emails to approval1 and admin
            subject_approval1 = "Pendaftaran Calon Anggota Baru - Membutuhkan Approval 1"
            body_approval1 = f"""
            <html>
            <body style="font-family: Arial, sans-serif;">
                <h3 style="color: #1e5a96;">Ada Pengajuan Pendaftaran Calon Anggota Baru!</h3>
                <p>Pengajuan baru yang membutuhkan review dan approval dari Anda:</p>
                <hr>
                <table style="border-collapse: collapse; width: 100%;">
                    <tr>
                        <td style="padding: 8px; font-weight: bold;">Nama Lengkap:</td>
                        <td style="padding: 8px;">{data['nama_lengkap']}</td>
                    </tr>
                    <tr style="background-color: #f5f5f5;">
                        <td style="padding: 8px; font-weight: bold;">Departemen:</td>
                        <td style="padding: 8px;">{data['asal_departemen']}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; font-weight: bold;">No. ID Karyawan:</td>
                        <td style="padding: 8px;">{data['no_id_karyawan']}</td>
                    </tr>
                    <tr style="background-color: #f5f5f5;">
                        <td style="padding: 8px; font-weight: bold;">Email:</td>
                        <td style="padding: 8px;">{data['alamat_email']}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; font-weight: bold;">No. Handphone:</td>
                        <td style="padding: 8px;">{data['nomor_handphone']}</td>
                    </tr>
                    <tr style="background-color: #f5f5f5;">
                        <td style="padding: 8px; font-weight: bold;">Tanggal Pendaftaran:</td>
                        <td style="padding: 8px;">{datetime.now().strftime('%d %B %Y %H:%M')}</td>
                    </tr>
                </table>
                <hr>
                <p style="color: #d32f2f;">Silakan login ke sistem untuk melakukan review dan approval:</p>
                <p><a href="http://127.0.0.1:5000/login" style="background-color: #1e5a96; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">Login ke Sistem</a></p>
                <hr>
                <p style="font-size: 12px; color: #999;">Email ini dikirim otomatis dari sistem Aplikasi Pendaftaran Koperasi. Jangan balas email ini.</p>
            </body>
            </html>
            """
            
            # Send to approval1 and admin
            try:
                send_email(app.config['APPROVAL1_EMAIL'], subject_approval1, body_approval1)
                send_email(app.config['ADMIN_EMAIL'], subject_approval1, body_approval1)
            except Exception as e:
                print(f"Error sending notification emails: {str(e)}")
            
            # Send confirmation email to applicant
            subject_applicant = "Pendaftaran Anda Telah Diterima - Kopkar SHTB"
            body_applicant = f"""
            <html>
            <body style="font-family: Arial, sans-serif;">
                <h3 style="color: #1e5a96;">Pendaftaran Anda Telah Diterima!</h3>
                <p>Terima kasih telah mendaftar sebagai calon anggota Koperasi Karyawan RS Siloam Hospitals TB Simatupang.</p>
                <hr>
                <p><strong>Data Pendaftaran Anda:</strong></p>
                <table style="border-collapse: collapse; width: 100%;">
                    <tr>
                        <td style="padding: 8px; font-weight: bold;">Nama:</td>
                        <td style="padding: 8px;">{data['nama_lengkap']}</td>
                    </tr>
                    <tr style="background-color: #f5f5f5;">
                        <td style="padding: 8px; font-weight: bold;">Departemen:</td>
                        <td style="padding: 8px;">{data['asal_departemen']}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; font-weight: bold;">Tanggal Pendaftaran:</td>
                        <td style="padding: 8px;">{datetime.now().strftime('%d %B %Y %H:%M')}</td>
                    </tr>
                </table>
                <hr>
                <p style="color: #1e5a96;"><strong>Status Proses:</strong></p>
                <p>Pengajuan Anda sedang dalam tahap <strong>Approval 1</strong>. Anda akan menerima notifikasi email untuk setiap perubahan status pengajuan.</p>
                <hr>
                <p style="font-size: 12px; color: #999;">Email ini dikirim otomatis dari sistem Aplikasi Pendaftaran Koperasi. Jangan balas email ini.</p>
            </body>
            </html>
            """
            
            try:
                send_email(data['alamat_email'], subject_applicant, body_applicant)
            except Exception as e:
                print(f"Error sending confirmation email to applicant: {str(e)}")
            
            flash('Pendaftaran berhasil dikirim. Status akan diinformasikan via email.', 'success')
            return redirect(url_for('index'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Terjadi kesalahan: {str(e)}', 'danger')
    
    return render_template('register.html', departemen_options=app.config['DEPARTEMEN_OPTIONS'])

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        user = User.query.filter_by(email=email).first()
        
        if user and user.check_password(password):
            login_user(user)
            app.logger.info(f'User {email} logged in successfully')
            next_page = request.args.get('next')
            return redirect(next_page or url_for('dashboard'))
        else:
            app.logger.warning(f'Failed login attempt for email: {email}')
            flash('Email atau password salah', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    app.logger.info(f'User {current_user.email} logged out')
    logout_user()
    return redirect(url_for('index'))

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        old_password = request.form.get('old_password', '')
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')
        
        if not current_user.check_password(old_password):
            flash('Password lama tidak sesuai', 'danger')
            return redirect(url_for('change_password'))
        
        if new_password != confirm_password:
            flash('Password baru tidak cocok', 'danger')
            return redirect(url_for('change_password'))
        
        if len(new_password) < 6:
            flash('Password minimal 6 karakter', 'danger')
            return redirect(url_for('change_password'))
        
        try:
            current_user.set_password(new_password)
            db.session.commit()
            flash('Password berhasil diubah', 'success')
            return redirect(url_for('dashboard'))
        except Exception as e:
            db.session.rollback()
            flash(f'Terjadi kesalahan: {str(e)}', 'danger')
    
    return render_template('change_password.html')

@app.route('/dashboard')
@login_required
def dashboard():
    if current_user.role == 'user':
        pendaftaran = Pendaftaran.query.filter_by(alamat_email=current_user.email).first()
        return render_template('dashboard.html', pendaftaran=pendaftaran, user=current_user)
    else:
        # For approval users and admin
        if current_user.role == 'approval1':
            pendaftaran_list = Pendaftaran.query.filter_by(status='pending').all()
        elif current_user.role == 'approval2':
            pendaftaran_list = Pendaftaran.query.filter_by(status='approved1').all()
        else:  # admin
            pendaftaran_list = Pendaftaran.query.all()
        
        return render_template('admin.html', pendaftaran_list=pendaftaran_list, user=current_user)

@app.route('/approve/<int:id>', methods=['POST'])
@login_required
def approve(id):
    app.logger.info(f'Approval action initiated by {current_user.email} for pendaftaran ID {id}')
    if current_user.role not in ['approval1', 'approval2', 'admin']:
        flash('Anda tidak memiliki akses', 'danger')
        return redirect(url_for('dashboard'))
    
    pendaftaran = Pendaftaran.query.get_or_404(id)
    action = request.form['action']
    notes = request.form.get('notes', '')
    
    if current_user.role == 'approval1':
        if action == 'approve':
            pendaftaran.status = 'approved1'
            pendaftaran.approval1_by = current_user.id
            pendaftaran.approval1_at = datetime.utcnow()
            pendaftaran.approval1_notes = notes
            
            # Send email to approval2
            subject = "Pengajuan Calon Anggota - Menunggu Approval 2"
            body = f"""
            <html>
            <body style="font-family: Arial, sans-serif;">
                <h3 style="color: #1e5a96;">Pengajuan Telah Disetujui Approval 1!</h3>
                <p>Pengajuan calon anggota telah disetujui oleh Approval 1 dan membutuhkan review dari Anda:</p>
                <hr>
                <table style="border-collapse: collapse; width: 100%;">
                    <tr>
                        <td style="padding: 8px; font-weight: bold;">Nama Lengkap:</td>
                        <td style="padding: 8px;">{pendaftaran.nama_lengkap}</td>
                    </tr>
                    <tr style="background-color: #f5f5f5;">
                        <td style="padding: 8px; font-weight: bold;">Departemen:</td>
                        <td style="padding: 8px;">{pendaftaran.asal_departemen}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; font-weight: bold;">No. ID Karyawan:</td>
                        <td style="padding: 8px;">{pendaftaran.no_id_karyawan}</td>
                    </tr>
                    <tr style="background-color: #f5f5f5;">
                        <td style="padding: 8px; font-weight: bold;">Disetujui Oleh:</td>
                        <td style="padding: 8px;">{current_user.email}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; font-weight: bold;">Catatan:</td>
                        <td style="padding: 8px;">{notes if notes else '-'}</td>
                    </tr>
                </table>
                <hr>
                <p style="color: #d32f2f;">Silakan login ke sistem untuk melakukan review dan approval final:</p>
                <p><a href="http://127.0.0.1:5000/login" style="background-color: #1e5a96; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">Login ke Sistem</a></p>
                <hr>
                <p style="font-size: 12px; color: #999;">Email ini dikirim otomatis dari sistem Aplikasi Pendaftaran Koperasi.</p>
            </body>
            </html>
            """
            
            try:
                send_email(app.config['APPROVAL2_EMAIL'], subject, body)
                send_email(app.config['ADMIN_EMAIL'], subject, body)  # Notify admin too
            except Exception as e:
                print(f"Error sending approval2 notification: {str(e)}")
            
            # Send notification to applicant
            subject_applicant = "Pengajuan Anda Diterima - Tahap Approval 2"
            body_applicant = f"""
            <html>
            <body style="font-family: Arial, sans-serif;">
                <h3 style="color: #1e5a96;">Pengajuan Anda Tahap Selanjutnya!</h3>
                <p>Pengajuan Anda telah disetujui pada tahap Approval 1 dan sedang dalam proses Approval 2.</p>
                <p>Anda akan menerima notifikasi email untuk setiap perkembangan proses pengajuan Anda.</p>
                <hr>
                <p style="color: #1e5a96;"><strong>Status Saat Ini:</strong> Approval 2 (Tahap Final)</p>
                <hr>
                <p style="font-size: 12px; color: #999;">Email ini dikirim otomatis dari sistem Aplikasi Pendaftaran Koperasi.</p>
            </body>
            </html>
            """
            
            try:
                send_email(pendaftaran.alamat_email, subject_applicant, body_applicant)
            except Exception as e:
                print(f"Error sending applicant notification: {str(e)}")
            
            flash('Approval 1 berhasil diberikan', 'success')
            
        elif action == 'reject':
            pendaftaran.status = 'rejected'
            pendaftaran.approval1_by = current_user.id
            pendaftaran.approval1_at = datetime.utcnow()
            pendaftaran.approval1_notes = notes
            
            # Send email to applicant
            subject_applicant = "Pengajuan Anda Ditolak - Kopkar SHTB"
            body_applicant = f"""
            <html>
            <body style="font-family: Arial, sans-serif;">
                <h3 style="color: #d32f2f;">Pengajuan Ditolak</h3>
                <p>Mohon maaf, pengajuan Anda sebagai calon anggota Koperasi Karyawan RS Siloam Hospitals TB Simatupang ditolak pada tahap Approval 1.</p>
                <hr>
                <p><strong>Alasan Penolakan:</strong></p>
                <p>{notes if notes else 'Tidak ada keterangan'}</p>
                <hr>
                <p>Silakan menghubungi administrator untuk informasi lebih lanjut atau melakukan pengajuan ulang.</p>
                <p>Email Admin: {app.config['ADMIN_EMAIL']}</p>
                <hr>
                <p style="font-size: 12px; color: #999;">Email ini dikirim otomatis dari sistem Aplikasi Pendaftaran Koperasi.</p>
            </body>
            </html>
            """
            
            # Send email to admin
            subject_admin = "Pengajuan Ditolak - Approval 1"
            body_admin = f"""
            <html>
            <body style="font-family: Arial, sans-serif;">
                <h3 style="color: #d32f2f;">Pengajuan Telah Ditolak oleh Approval 1</h3>
                <table style="border-collapse: collapse; width: 100%;">
                    <tr>
                        <td style="padding: 8px; font-weight: bold;">Nama:</td>
                        <td style="padding: 8px;">{pendaftaran.nama_lengkap}</td>
                    </tr>
                    <tr style="background-color: #f5f5f5;">
                        <td style="padding: 8px; font-weight: bold;">Departemen:</td>
                        <td style="padding: 8px;">{pendaftaran.asal_departemen}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; font-weight: bold;">Alasan:</td>
                        <td style="padding: 8px;">{notes if notes else '-'}</td>
                    </tr>
                    <tr style="background-color: #f5f5f5;">
                        <td style="padding: 8px; font-weight: bold;">Ditolak Oleh:</td>
                        <td style="padding: 8px;">{current_user.email}</td>
                    </tr>
                </table>
                <hr>
                <p style="font-size: 12px; color: #999;">Email ini dikirim otomatis dari sistem Aplikasi Pendaftaran Koperasi.</p>
            </body>
            </html>
            """
            
            try:
                send_email(pendaftaran.alamat_email, subject_applicant, body_applicant)
                send_email(app.config['ADMIN_EMAIL'], subject_admin, body_admin)
            except Exception as e:
                print(f"Error sending rejection emails: {str(e)}")
            
            flash('Pendaftaran ditolak', 'warning')
    
    elif current_user.role == 'approval2':
        if action == 'approve':
            pendaftaran.status = 'approved2'
            pendaftaran.approval2_by = current_user.id
            pendaftaran.approval2_at = datetime.utcnow()
            pendaftaran.approval2_notes = notes
            
            # Send email to admin with HTML formatting
            subject_admin = f"✓ Pendaftaran Calon Anggota Disetujui Lengkap - {pendaftaran.nama_lengkap}"
            body_admin = f"""
            <html>
            <body style="font-family: Arial, sans-serif;">
                <h3 style="color: #1e5a96; border-bottom: 2px solid #1e5a96; padding-bottom: 10px;">PERSETUJUAN FINAL - APPROVAL 2</h3>
                <p>Pendaftaran calon anggota telah <strong style="color: #2e7d32;">DISETUJUI LENGKAP</strong> melalui tahap Approval 2.</p>
                
                <h4 style="color: #1e5a96; margin-top: 20px;">Informasi Calon Anggota:</h4>
                <table style="border-collapse: collapse; width: 100%; margin-bottom: 20px;">
                    <tr>
                        <td style="padding: 8px; font-weight: bold; width: 30%;">Nama Lengkap:</td>
                        <td style="padding: 8px;">{pendaftaran.nama_lengkap}</td>
                    </tr>
                    <tr style="background-color: #f5f5f5;">
                        <td style="padding: 8px; font-weight: bold;">No. ID Karyawan:</td>
                        <td style="padding: 8px;">{pendaftaran.no_id_karyawan}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; font-weight: bold;">Departemen:</td>
                        <td style="padding: 8px;">{pendaftaran.asal_departemen}</td>
                    </tr>
                    <tr style="background-color: #f5f5f5;">
                        <td style="padding: 8px; font-weight: bold;">Email:</td>
                        <td style="padding: 8px;">{pendaftaran.alamat_email}</td>
                    </tr>
                </table>
                
                <h4 style="color: #1e5a96; margin-top: 20px;">Status Persetujuan:</h4>
                <table style="border-collapse: collapse; width: 100%; margin-bottom: 20px;">
                    <tr>
                        <td style="padding: 8px; font-weight: bold; width: 30%;">Approval 1:</td>
                        <td style="padding: 8px; color: #2e7d32;"><strong>✓ DISETUJUI</strong></td>
                    </tr>
                    <tr style="background-color: #f5f5f5;">
                        <td style="padding: 8px; font-weight: bold;">Approval 1 Oleh:</td>
                        <td style="padding: 8px;">{User.query.get(pendaftaran.approval1_by).email if pendaftaran.approval1_by else '-'}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; font-weight: bold;">Approval 2:</td>
                        <td style="padding: 8px; color: #2e7d32;"><strong>✓ DISETUJUI</strong></td>
                    </tr>
                    <tr style="background-color: #f5f5f5;">
                        <td style="padding: 8px; font-weight: bold;">Approval 2 Oleh:</td>
                        <td style="padding: 8px;">{current_user.email}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; font-weight: bold;">Catatan Approval 2:</td>
                        <td style="padding: 8px;">{notes if notes else '-'}</td>
                    </tr>
                </table>
                
                <p style="text-align: center; margin-top: 20px;">
                    <a href="http://127.0.0.1:5000/login" style="display: inline-block; background-color: #1e5a96; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">Lihat di Sistem</a>
                </p>
                <hr style="margin: 20px 0; border: none; border-top: 1px solid #ddd;">
                <p style="font-size: 12px; color: #999;">Email ini dikirim otomatis dari sistem Aplikasi Pendaftaran Koperasi.</p>
            </body>
            </html>
            """
            
            try:
                send_email(app.config['ADMIN_EMAIL'], subject_admin, body_admin)
            except Exception as e:
                print(f"Error sending approval2 admin email: {str(e)}")
            
            # Send email to applicant
            subject_applicant = "✓ Pendaftaran Anda Disetujui Lengkap - Selamat Bergabung!"
            body_applicant = f"""
            <html>
            <body style="font-family: Arial, sans-serif;">
                <h3 style="color: #2e7d32; border-bottom: 2px solid #2e7d32; padding-bottom: 10px;">SELAMAT! PENDAFTARAN ANDA DISETUJUI!</h3>
                <p>Dengan ini kami informasikan bahwa pendaftaran Anda sebagai calon anggota Koperasi Karyawan RS Siloam Hospitals TB Simatupang telah <strong style="color: #2e7d32;">DISETUJUI SECARA LENGKAP</strong>.</p>
                
                <h4 style="color: #1e5a96; margin-top: 20px;">Informasi Persetujuan:</h4>
                <table style="border-collapse: collapse; width: 100%; margin-bottom: 20px;">
                    <tr>
                        <td style="padding: 8px; font-weight: bold; width: 40%;">Nama Lengkap:</td>
                        <td style="padding: 8px;">{pendaftaran.nama_lengkap}</td>
                    </tr>
                    <tr style="background-color: #f5f5f5;">
                        <td style="padding: 8px; font-weight: bold;">No. ID Karyawan:</td>
                        <td style="padding: 8px;">{pendaftaran.no_id_karyawan}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; font-weight: bold;">Departemen:</td>
                        <td style="padding: 8px;">{pendaftaran.asal_departemen}</td>
                    </tr>
                    <tr style="background-color: #f5f5f5;">
                        <td style="padding: 8px; font-weight: bold;">Status:</td>
                        <td style="padding: 8px; color: #2e7d32;"><strong>ANGGOTA DISETUJUI</strong></td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; font-weight: bold;">Tanggal Disetujui:</td>
                        <td style="padding: 8px;">{datetime.utcnow().strftime('%d-%m-%Y %H:%M:%S')}</td>
                    </tr>
                </table>
                
                <p style="margin-top: 20px; line-height: 1.6;">
                    Anda sekarang resmi menjadi anggota <strong>Koperasi Karyawan RS Siloam Hospitals TB Simatupang</strong>.
                    <br><br>
                    Silakan menunggu informasi lebih lanjut dari administrator mengenai prosedur aktivasi akun dan hak-hak sebagai anggota koperasi.
                </p>
                
                <p style="text-align: center; margin-top: 20px;">
                    <a href="http://127.0.0.1:5000/login" style="display: inline-block; background-color: #2e7d32; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">Login ke Sistem</a>
                </p>
                <hr style="margin: 20px 0; border: none; border-top: 1px solid #ddd;">
                <p style="font-size: 12px; color: #999;">Email ini dikirim otomatis dari sistem Aplikasi Pendaftaran Koperasi.</p>
            </body>
            </html>
            """
            
            try:
                send_email(pendaftaran.alamat_email, subject_applicant, body_applicant)
            except Exception as e:
                print(f"Error sending approval2 applicant email: {str(e)}")
            
            flash('Approval 2 berhasil diberikan', 'success')
        
        elif action == 'reject':
            pendaftaran.status = 'rejected'
            pendaftaran.approval2_by = current_user.id
            pendaftaran.approval2_at = datetime.utcnow()
            pendaftaran.approval2_notes = notes
            
            # Send email to applicant
            subject_applicant = "✗ Pendaftaran Anda Ditolak oleh Approval 2"
            body_applicant = f"""
            <html>
            <body style="font-family: Arial, sans-serif;">
                <h3 style="color: #d32f2f; border-bottom: 2px solid #d32f2f; padding-bottom: 10px;">PENDAFTARAN DITOLAK</h3>
                <p>Mohon maaf, pendaftaran Anda sebagai calon anggota Koperasi telah <strong style="color: #d32f2f;">DITOLAK</strong> pada tahap Approval 2.</p>
                
                <h4 style="color: #1e5a96; margin-top: 20px;">Informasi Penolakan:</h4>
                <table style="border-collapse: collapse; width: 100%; margin-bottom: 20px;">
                    <tr>
                        <td style="padding: 8px; font-weight: bold; width: 40%;">Nama Lengkap:</td>
                        <td style="padding: 8px;">{pendaftaran.nama_lengkap}</td>
                    </tr>
                    <tr style="background-color: #f5f5f5;">
                        <td style="padding: 8px; font-weight: bold;">No. ID Karyawan:</td>
                        <td style="padding: 8px;">{pendaftaran.no_id_karyawan}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; font-weight: bold;">Alasan Penolakan:</td>
                        <td style="padding: 8px; color: #d32f2f;">{notes if notes else 'Tidak ada keterangan'}</td>
                    </tr>
                    <tr style="background-color: #f5f5f5;">
                        <td style="padding: 8px; font-weight: bold;">Ditolak Oleh:</td>
                        <td style="padding: 8px;">{current_user.email}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; font-weight: bold;">Tanggal Penolakan:</td>
                        <td style="padding: 8px;">{datetime.utcnow().strftime('%d-%m-%Y %H:%M:%S')}</td>
                    </tr>
                </table>
                
                <p style="margin-top: 20px; line-height: 1.6;">
                    Silakan hubungi administrator untuk informasi lebih lanjut mengenai alasan penolakan dan prosedur berikutnya.
                </p>
                
                <p style="text-align: center; margin-top: 20px;">
                    <a href="http://127.0.0.1:5000/login" style="display: inline-block; background-color: #1e5a96; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">Hubungi Administrator</a>
                </p>
                <hr style="margin: 20px 0; border: none; border-top: 1px solid #ddd;">
                <p style="font-size: 12px; color: #999;">Email ini dikirim otomatis dari sistem Aplikasi Pendaftaran Koperasi.</p>
            </body>
            </html>
            """
            
            try:
                send_email(pendaftaran.alamat_email, subject_applicant, body_applicant)
            except Exception as e:
                print(f"Error sending approval2 rejection applicant email: {str(e)}")
            
            # Send email to admin
            subject_admin = f"✗ Pendaftaran Calon Anggota Ditolak oleh Approval 2 - {pendaftaran.nama_lengkap}"
            body_admin = f"""
            <html>
            <body style="font-family: Arial, sans-serif;">
                <h3 style="color: #d32f2f; border-bottom: 2px solid #d32f2f; padding-bottom: 10px;">PENDAFTARAN DITOLAK - APPROVAL 2</h3>
                <p>Pendaftaran calon anggota telah <strong style="color: #d32f2f;">DITOLAK</strong> oleh Approval 2.</p>
                
                <h4 style="color: #1e5a96; margin-top: 20px;">Informasi Calon Anggota:</h4>
                <table style="border-collapse: collapse; width: 100%; margin-bottom: 20px;">
                    <tr>
                        <td style="padding: 8px; font-weight: bold; width: 30%;">Nama Lengkap:</td>
                        <td style="padding: 8px;">{pendaftaran.nama_lengkap}</td>
                    </tr>
                    <tr style="background-color: #f5f5f5;">
                        <td style="padding: 8px; font-weight: bold;">No. ID Karyawan:</td>
                        <td style="padding: 8px;">{pendaftaran.no_id_karyawan}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; font-weight: bold;">Departemen:</td>
                        <td style="padding: 8px;">{pendaftaran.asal_departemen}</td>
                    </tr>
                    <tr style="background-color: #f5f5f5;">
                        <td style="padding: 8px; font-weight: bold;">Email:</td>
                        <td style="padding: 8px;">{pendaftaran.alamat_email}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; font-weight: bold;">Alasan Penolakan:</td>
                        <td style="padding: 8px; color: #d32f2f;">{notes if notes else 'Tidak ada keterangan'}</td>
                    </tr>
                    <tr style="background-color: #f5f5f5;">
                        <td style="padding: 8px; font-weight: bold;">Ditolak Oleh:</td>
                        <td style="padding: 8px;">{current_user.email}</td>
                    </tr>
                </table>
                
                <hr style="margin: 20px 0; border: none; border-top: 1px solid #ddd;">
                <p style="font-size: 12px; color: #999;">Email ini dikirim otomatis dari sistem Aplikasi Pendaftaran Koperasi.</p>
            </body>
            </html>
            """
            
            try:
                send_email(app.config['ADMIN_EMAIL'], subject_admin, body_admin)
            except Exception as e:
                print(f"Error sending approval2 rejection admin email: {str(e)}")
            
            flash('Pendaftaran ditolak', 'warning')
    
    db.session.commit()
    app.logger.info(f'Approval action completed by {current_user.email} for pendaftaran ID {id}: {action}')
    return redirect(url_for('dashboard'))

@app.route('/delete-pendaftaran/<int:pendaftaran_id>', methods=['POST'])
@login_required
def delete_pendaftaran(pendaftaran_id):
    """Delete pendaftaran dan file uploads - hanya untuk admin"""
    # Check authorization - only admin can delete
    if current_user.role != 'admin':
        flash('Anda tidak memiliki akses untuk menghapus pendaftaran', 'danger')
        return redirect(url_for('dashboard'))
    
    pendaftaran = Pendaftaran.query.get_or_404(pendaftaran_id)
    
    # Delete files from disk
    file_fields = ['upload_ktp', 'upload_id_karyawan', 'upload_pas_foto', 'upload_buku_tabungan']
    folder_map = {
        'upload_ktp': 'ktp',
        'upload_id_karyawan': 'id_karyawan',
        'upload_pas_foto': 'pas_foto',
        'upload_buku_tabungan': 'buku_tabungan'
    }
    
    for field in file_fields:
        filename = getattr(pendaftaran, field, None)
        if filename:
            folder = folder_map[field]
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], folder, filename)
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
            except Exception as e:
                print(f"Error deleting file {file_path}: {str(e)}")
    
    # Delete from database
    try:
        nama_lengkap = pendaftaran.nama_lengkap
        nomor_ktp = pendaftaran.nomor_ktp
        db.session.delete(pendaftaran)
        db.session.commit()
        flash(f'Pendaftaran {nama_lengkap} (KTP: {nomor_ktp}) berhasil dihapus. Nomor KTP dapat melakukan pendaftaran ulang.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Terjadi kesalahan saat menghapus pendaftaran: {str(e)}', 'danger')
    
    return redirect(url_for('dashboard'))

@app.route('/download/<int:pendaftaran_id>/<file_type>')
@login_required
def download_file(pendaftaran_id, file_type):
    """Download file uploaded by calon anggota"""
    # Check authorization - only approval1, approval2, and admin can download
    if current_user.role not in ['approval1', 'approval2', 'admin']:
        flash('Anda tidak memiliki akses untuk mengunduh file', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get the pendaftaran record
    pendaftaran = Pendaftaran.query.get_or_404(pendaftaran_id)
    
    # Map file_type to database field
    file_mapping = {
        'ktp': 'upload_ktp',
        'id_karyawan': 'upload_id_karyawan',
        'pas_foto': 'upload_pas_foto',
        'buku_tabungan': 'upload_buku_tabungan'
    }
    
    if file_type not in file_mapping:
        flash('Tipe file tidak valid', 'danger')
        return redirect(request.referrer or url_for('dashboard'))
    
    # Get the file name from database
    db_field = file_mapping[file_type]
    filename = getattr(pendaftaran, db_field, None)
    
    if not filename:
        flash('File tidak ditemukan', 'danger')
        return redirect(request.referrer or url_for('dashboard'))
    
    # Build full file path
    folder_map = {
        'ktp': 'ktp',
        'id_karyawan': 'id_karyawan',
        'pas_foto': 'pas_foto',
        'buku_tabungan': 'buku_tabungan'
    }
    
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], folder_map[file_type], filename)
    
    # Check if file exists
    if not os.path.exists(file_path):
        flash('File tidak ditemukan di server', 'danger')
        return redirect(request.referrer or url_for('dashboard'))
    
    try:
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        flash(f'Terjadi kesalahan saat mengunduh file: {str(e)}', 'danger')
        return redirect(request.referrer or url_for('dashboard'))

@app.route('/preview/<int:pendaftaran_id>/<file_type>')
@login_required
def preview_file(pendaftaran_id, file_type):
    """Preview file uploaded by calon anggota"""
    # Check authorization - only approval1, approval2, and admin can preview
    if current_user.role not in ['approval1', 'approval2', 'admin']:
        flash('Anda tidak memiliki akses untuk melihat file', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get the pendaftaran record
    pendaftaran = Pendaftaran.query.get_or_404(pendaftaran_id)
    
    # Map file_type to database field
    file_mapping = {
        'ktp': 'upload_ktp',
        'id_karyawan': 'upload_id_karyawan',
        'pas_foto': 'upload_pas_foto',
        'buku_tabungan': 'upload_buku_tabungan'
    }
    
    if file_type not in file_mapping:
        flash('Tipe file tidak valid', 'danger')
        return redirect(request.referrer or url_for('dashboard'))
    
    # Get the file name from database
    db_field = file_mapping[file_type]
    filename = getattr(pendaftaran, db_field, None)
    
    if not filename:
        flash('File tidak ditemukan', 'danger')
        return redirect(request.referrer or url_for('dashboard'))
    
    # Build full file path
    folder_map = {
        'ktp': 'ktp',
        'id_karyawan': 'id_karyawan',
        'pas_foto': 'pas_foto',
        'buku_tabungan': 'buku_tabungan'
    }
    
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], folder_map[file_type], filename)
    
    # Check if file exists
    if not os.path.exists(file_path):
        flash('File tidak ditemukan di server', 'danger')
        return redirect(request.referrer or url_for('dashboard'))
    
    try:
        # Return file inline for preview (not as attachment)
        return send_file(
            file_path,
            as_attachment=False
        )
    except Exception as e:
        flash(f'Terjadi kesalahan saat membuka file: {str(e)}', 'danger')
        return redirect(request.referrer or url_for('dashboard'))

@app.route('/pendaftaran/<int:pendaftaran_id>')
@login_required
def detail_pendaftaran(pendaftaran_id):
    """View detail pendaftaran dengan file"""
    # Check authorization
    if current_user.role not in ['approval1', 'approval2', 'admin']:
        flash('Anda tidak memiliki akses', 'danger')
        return redirect(url_for('dashboard'))
    
    pendaftaran = Pendaftaran.query.get_or_404(pendaftaran_id)
    
    return render_template('detail_pendaftaran.html', pendaftaran=pendaftaran, user=current_user)

@app.route('/report')
@login_required
def report():
    # Only admin and approval users can view report
    if current_user.role not in ['admin', 'approval1', 'approval2']:
        flash('Anda tidak memiliki akses', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get filter from request
    status_filter = request.args.get('status', 'all')
    approval1_filter = request.args.get('approval1', 'all')
    approval2_filter = request.args.get('approval2', 'all')
    search = request.args.get('search', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Build query
    query = Pendaftaran.query
    
    if status_filter != 'all':
        query = query.filter_by(status=status_filter)
    
    if approval1_filter != 'all':
        if approval1_filter == 'approved':
            query = query.filter(Pendaftaran.approval1_at.isnot(None))
        elif approval1_filter == 'pending':
            query = query.filter(Pendaftaran.approval1_at.is_(None))
    
    if approval2_filter != 'all':
        if approval2_filter == 'approved':
            query = query.filter(Pendaftaran.approval2_at.isnot(None))
        elif approval2_filter == 'pending':
            query = query.filter(Pendaftaran.approval2_at.is_(None))
    
    if search:
        query = query.filter(
            (Pendaftaran.nama_lengkap.ilike(f'%{search}%')) |
            (Pendaftaran.nomor_ktp.ilike(f'%{search}%')) |
            (Pendaftaran.no_id_karyawan.ilike(f'%{search}%')) |
            (Pendaftaran.alamat_email.ilike(f'%{search}%'))
        )
    
    # Date range filtering
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(Pendaftaran.created_at >= date_from_obj)
        except:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            # Add 1 day to include entire day
            date_to_obj = date_to_obj.replace(hour=23, minute=59, second=59)
            query = query.filter(Pendaftaran.created_at <= date_to_obj)
        except:
            pass
    
    pendaftaran_list = query.order_by(Pendaftaran.created_at.desc()).all()
    
    # Get statistics
    stats = {
        'total': Pendaftaran.query.count(),
        'pending': Pendaftaran.query.filter_by(status='pending').count(),
        'approved1': Pendaftaran.query.filter_by(status='approved1').count(),
        'approved2': Pendaftaran.query.filter_by(status='approved2').count(),
        'final_approved': Pendaftaran.query.filter(Pendaftaran.approval2_at.isnot(None)).count(),
        'rejected': Pendaftaran.query.filter_by(status='rejected').count(),
    }
    
    return render_template('report.html', 
                         pendaftaran_list=pendaftaran_list, 
                         status_filter=status_filter,
                         approval1_filter=approval1_filter,
                         approval2_filter=approval2_filter,
                         search=search,
                         date_from=date_from,
                         date_to=date_to,
                         stats=stats)

@app.route('/report/export')
@login_required
def export_report():
    # Only admin can export
    if current_user.role not in ['admin', 'approval1', 'approval2']:
        flash('Anda tidak memiliki akses', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get all pendaftaran data
    status_filter = request.args.get('status', 'all')
    approval1_filter = request.args.get('approval1', 'all')
    approval2_filter = request.args.get('approval2', 'all')
    search = request.args.get('search', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    query = Pendaftaran.query
    
    if status_filter != 'all':
        query = query.filter_by(status=status_filter)
    
    if approval1_filter != 'all':
        if approval1_filter == 'approved':
            query = query.filter(Pendaftaran.approval1_at.isnot(None))
        elif approval1_filter == 'pending':
            query = query.filter(Pendaftaran.approval1_at.is_(None))
    
    if approval2_filter != 'all':
        if approval2_filter == 'approved':
            query = query.filter(Pendaftaran.approval2_at.isnot(None))
        elif approval2_filter == 'pending':
            query = query.filter(Pendaftaran.approval2_at.is_(None))
    
    if search:
        query = query.filter(
            (Pendaftaran.nama_lengkap.ilike(f'%{search}%')) |
            (Pendaftaran.nomor_ktp.ilike(f'%{search}%')) |
            (Pendaftaran.no_id_karyawan.ilike(f'%{search}%')) |
            (Pendaftaran.alamat_email.ilike(f'%{search}%'))
        )
    
    # Date range filtering
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(Pendaftaran.created_at >= date_from_obj)
        except:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            date_to_obj = date_to_obj.replace(hour=23, minute=59, second=59)
            query = query.filter(Pendaftaran.created_at <= date_to_obj)
        except:
            pass
    
    pendaftaran_list = query.order_by(Pendaftaran.created_at.desc()).all()
    
    # Create CSV
    output = StringIO()
    writer = csv.writer(output)
    
    # Write header with approval dates
    writer.writerow([
        'No', 'Nama Lengkap', 'Tempat Lahir', 'Tanggal Lahir', 'Alamat Rumah',
        'No. Handphone', 'Email', 'No. KTP', 'No. ID Karyawan', 'Departemen',
        'No. Rekening', 'Bank', 'Status', 'Tgl Daftar', 'Approval 1', 'Tgl Approval 1',
        'Approval 2', 'Tgl Approval 2'
    ])
    
    # Write data
    for idx, p in enumerate(pendaftaran_list, 1):
        approval1_user = User.query.get(p.approval1_by).email if p.approval1_by else '-'
        approval2_user = User.query.get(p.approval2_by).email if p.approval2_by else '-'
        approval1_date = p.approval1_at.strftime('%Y-%m-%d %H:%M') if p.approval1_at else '-'
        approval2_date = p.approval2_at.strftime('%Y-%m-%d %H:%M') if p.approval2_at else '-'
        
        # Format nomor handphone dan nomor rekening agar leading zero tetap muncul
        nomor_handphone = f"'{p.nomor_handphone}" if p.nomor_handphone else ''
        nomor_rekening = f"'{p.nomor_rekening}" if p.nomor_rekening else ''
        
        writer.writerow([
            idx,
            p.nama_lengkap,
            p.tempat_lahir,
            p.tanggal_lahir.strftime('%Y-%m-%d') if p.tanggal_lahir else '',
            p.alamat_rumah,
            nomor_handphone,
            p.alamat_email,
            p.nomor_ktp,
            p.no_id_karyawan,
            p.asal_departemen,
            nomor_rekening,
            p.nama_bank,
            p.status,
            p.created_at.strftime('%Y-%m-%d %H:%M') if p.created_at else '',
            approval1_user,
            approval1_date,
            approval2_user,
            approval2_date
        ])
    
    # Create bytes response
    output.seek(0)
    mem = BytesIO()
    mem.write(output.getvalue().encode('utf-8-sig'))  # UTF-8 with BOM for Excel
    mem.seek(0)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'laporan_pendaftaran_{timestamp}.csv'
    
    return send_file(
        mem,
        mimetype='text/csv',
        as_attachment=True,
        download_name=filename
    )

@app.route('/dokter')
def dokter_list():
    # User biasa dan approved members bisa lihat dokter
    dokter_list_data = Dokter.query.filter_by(is_active=True).all()
    return render_template('dokter_list.html', dokter_list=dokter_list_data)

@app.route('/dokter/potong/<int:dokter_id>', methods=['POST'])
@login_required
def potong_kupon(dokter_id):
    dokter = Dokter.query.get_or_404(dokter_id)
    
    if dokter.jumlah_kupon <= 0:
        flash('Kupon dokter sudah habis', 'danger')
        return redirect(url_for('dokter_list'))
    
    try:
        jumlah = int(request.form.get('jumlah_kupon', 1))
        catatan = request.form.get('catatan', '')
        
        if jumlah <= 0:
            flash('Jumlah kupon harus lebih dari 0', 'danger')
            return redirect(url_for('dokter_list'))
        
        if jumlah > dokter.jumlah_kupon:
            flash(f'Kupon tidak cukup. Tersedia hanya {dokter.jumlah_kupon}', 'danger')
            return redirect(url_for('dokter_list'))
        
        # Buat record penggunaan kupon
        penggunaan = PenggunaanKupon(
            dokter_id=dokter_id,
            user_email=current_user.email,
            jumlah_kupon_digunakan=jumlah,
            catatan=catatan
        )
        db.session.add(penggunaan)
        
        # Kurangi jumlah kupon dokter
        dokter.jumlah_kupon -= jumlah
        
        db.session.commit()
        
        flash(f'Berhasil memotong {jumlah} kupon dari Dr. {dokter.nama_dokter}', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Terjadi kesalahan: {str(e)}', 'danger')
    
    return redirect(url_for('dokter_list'))

@app.route('/dokter/riwayat')
@login_required
def riwayat_kupon():
    riwayat = PenggunaanKupon.query.filter_by(user_email=current_user.email)\
        .order_by(PenggunaanKupon.tanggal_penggunaan.desc()).all()
    return render_template('riwayat_kupon.html', riwayat=riwayat)

@app.route('/admin/dokter', methods=['GET', 'POST'])
@login_required
def admin_dokter():
    if current_user.role != 'admin':
        flash('Anda tidak memiliki akses', 'danger')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'add':
            try:
                # Get current month for periode
                current_month = datetime.now().strftime('%Y-%m')
                
                dokter = Dokter(
                    nama_dokter=request.form['nama_dokter'],
                    nomor_kontak=request.form['nomor_kontak'],
                    jumlah_kupon=int(request.form['jumlah_kupon']),
                    kupon_awal=int(request.form['jumlah_kupon']),
                    periode_berlaku_bulan=current_month,
                    keterangan=request.form.get('keterangan', '')
                )
                db.session.add(dokter)
                db.session.commit()
                flash('Dokter berhasil ditambahkan', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Terjadi kesalahan: {str(e)}', 'danger')
        
        elif action == 'update':
            try:
                dokter_id = int(request.form['dokter_id'])
                dokter = Dokter.query.get_or_404(dokter_id)
                
                # Check if periode changed (month changed), reset kupon
                current_month = datetime.now().strftime('%Y-%m')
                if dokter.periode_berlaku_bulan != current_month:
                    # Month changed, reset kupon
                    jumlah_baru = int(request.form['jumlah_kupon'])
                    dokter.kupon_awal = jumlah_baru
                    dokter.jumlah_kupon = jumlah_baru
                    dokter.periode_berlaku_bulan = current_month
                else:
                    # Same month, only update kupon if needed
                    dokter.jumlah_kupon = int(request.form['jumlah_kupon'])
                
                dokter.nama_dokter = request.form['nama_dokter']
                dokter.nomor_kontak = request.form['nomor_kontak']
                dokter.keterangan = request.form.get('keterangan', '')
                
                db.session.commit()
                flash('Dokter berhasil diupdate', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Terjadi kesalahan: {str(e)}', 'danger')
        
        elif action == 'delete':
            try:
                dokter_id = int(request.form['dokter_id'])
                dokter = Dokter.query.get_or_404(dokter_id)
                dokter.is_active = False
                db.session.commit()
                flash('Dokter berhasil dihapus', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Terjadi kesalahan: {str(e)}', 'danger')
        
        return redirect(url_for('admin_dokter'))
    
    # Check and reset kupon for each dokter if month changed
    current_month = datetime.now().strftime('%Y-%m')
    for dokter in Dokter.query.all():
        if dokter.periode_berlaku_bulan and dokter.periode_berlaku_bulan != current_month:
            dokter.jumlah_kupon = dokter.kupon_awal
            dokter.periode_berlaku_bulan = current_month
    db.session.commit()
    
    dokter_list_data = Dokter.query.all()
    return render_template('admin_dokter.html', dokter_list=dokter_list_data)

@app.route('/admin/dokter-upload', methods=['POST'])
@login_required
def admin_dokter_upload():
    """Upload dokter data dari file Excel"""
    if current_user.role != 'admin':
        flash('Anda tidak memiliki akses', 'danger')
        return redirect(url_for('dashboard'))
    
    try:
        # Check if file was uploaded
        if 'excel_file' not in request.files:
            flash('Silakan pilih file Excel', 'danger')
            return redirect(url_for('admin_dokter'))
        
        file = request.files['excel_file']
        
        if file.filename == '':
            flash('Silakan pilih file Excel', 'danger')
            return redirect(url_for('admin_dokter'))
        
        # Check file extension
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Hanya format Excel (.xlsx, .xls) yang diizinkan', 'danger')
            return redirect(url_for('admin_dokter'))
        
        # Save file temporarily
        filename = secure_filename(file.filename)
        temp_path = os.path.join(tempfile.gettempdir(), filename)
        file.save(temp_path)
        
        # Load workbook
        workbook = load_workbook(temp_path)
        worksheet = workbook.active
        
        # Track results
        added_count = 0
        error_count = 0
        errors = []
        
        current_month = datetime.now().strftime('%Y-%m')
        
        # Skip header row, start from row 2
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Extract data from row
                nama_dokter = row[0]
                nomor_kontak = row[1]
                jumlah_kupon = row[2]
                keterangan = row[3] if len(row) > 3 else ''
                
                # Validasi data
                if not nama_dokter or not nomor_kontak or jumlah_kupon is None:
                    errors.append(f"Baris {row_num}: Nama dokter, No. kontak, dan Jumlah kupon harus diisi")
                    error_count += 1
                    continue
                
                # Convert to appropriate types
                nama_dokter = str(nama_dokter).strip()
                nomor_kontak = str(nomor_kontak).strip()
                
                try:
                    jumlah_kupon = int(float(jumlah_kupon))
                except (ValueError, TypeError):
                    errors.append(f"Baris {row_num}: Jumlah kupon harus berupa angka")
                    error_count += 1
                    continue
                
                keterangan = str(keterangan).strip() if keterangan else ''
                
                # Check if dokter already exists (by name and contact)
                existing = Dokter.query.filter_by(
                    nama_dokter=nama_dokter, 
                    nomor_kontak=nomor_kontak
                ).first()
                
                if existing:
                    errors.append(f"Baris {row_num}: Dokter '{nama_dokter}' dengan no. kontak '{nomor_kontak}' sudah terdaftar")
                    error_count += 1
                    continue
                
                # Create new dokter
                dokter = Dokter(
                    nama_dokter=nama_dokter,
                    nomor_kontak=nomor_kontak,
                    spesialisasi='-',  # Default value
                    jumlah_kupon=jumlah_kupon,
                    kupon_awal=jumlah_kupon,
                    periode_berlaku_bulan=current_month,
                    keterangan=keterangan,
                    is_active=True
                )
                
                db.session.add(dokter)
                added_count += 1
                
            except Exception as e:
                errors.append(f"Baris {row_num}: {str(e)}")
                error_count += 1
        
        # Commit semua perubahan
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal menyimpan data: {str(e)}', 'danger')
            return redirect(url_for('admin_dokter'))
        finally:
            # Delete temporary file
            if os.path.exists(temp_path):
                os.remove(temp_path)
        
        # Prepare success message
        if added_count > 0:
            flash(f'✓ Berhasil menambahkan {added_count} dokter', 'success')
        
        if error_count > 0:
            error_msg = f'⚠ {error_count} baris gagal ditambahkan:\n' + '\n'.join(errors[:5])
            if len(errors) > 5:
                error_msg += f'\n... dan {len(errors) - 5} error lainnya'
            flash(error_msg, 'warning')
        
        if added_count == 0 and error_count > 0:
            return redirect(url_for('admin_dokter'))
            
    except Exception as e:
        db.session.rollback()
        flash(f'Terjadi kesalahan saat memproses file: {str(e)}', 'danger')
    
    return redirect(url_for('admin_dokter'))

@app.route('/kasir/dokter', methods=['GET', 'POST'])
@login_required
def kasir_dokter():
    # Only kasir dapat akses
    if current_user.role != 'kasir':
        flash('Anda tidak memiliki akses', 'danger')
        return redirect(url_for('index'))
    
    search = request.args.get('search', '')
    
    # Build query
    query = Dokter.query.filter_by(is_active=True)
    
    if search:
        query = query.filter(
            (Dokter.nama_dokter.ilike(f'%{search}%')) |
            (Dokter.spesialisasi.ilike(f'%{search}%'))
        )
    
    dokter_list_data = query.all()
    
    return render_template('kasir_dokter.html', 
                         dokter_list=dokter_list_data, 
                         search=search)

@app.route('/kasir/dokter/potong/<int:dokter_id>', methods=['POST'])
@login_required
def kasir_potong_kupon(dokter_id):
    # Only kasir dapat akses
    if current_user.role != 'kasir':
        flash('Anda tidak memiliki akses', 'danger')
        return redirect(url_for('index'))
    
    dokter = Dokter.query.get_or_404(dokter_id)
    
    if dokter.jumlah_kupon <= 0:
        flash('Kupon dokter sudah habis', 'danger')
        return redirect(url_for('kasir_dokter'))
    
    try:
        jumlah = int(request.form.get('jumlah_kupon', 1))
        nama_member = request.form.get('nama_member', 'Unknown')
        catatan = request.form.get('catatan', '')
        
        if jumlah <= 0:
            flash('Jumlah kupon harus lebih dari 0', 'danger')
            return redirect(url_for('kasir_dokter'))
        
        if jumlah > dokter.jumlah_kupon:
            flash(f'Kupon tidak cukup. Tersedia hanya {dokter.jumlah_kupon}', 'danger')
            return redirect(url_for('kasir_dokter'))
        
        # Buat record penggunaan kupon
        penggunaan = PenggunaanKupon(
            dokter_id=dokter_id,
            user_email=nama_member,  # Store member name di user_email untuk kasir
            jumlah_kupon_digunakan=jumlah,
            catatan=catatan
        )
        db.session.add(penggunaan)
        
        # Kurangi jumlah kupon dokter
        dokter.jumlah_kupon -= jumlah
        
        db.session.commit()
        
        flash(f'Berhasil memotong {jumlah} kupon untuk {nama_member} dari Dr. {dokter.nama_dokter}', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Terjadi kesalahan: {str(e)}', 'danger')
    
    return redirect(url_for('kasir_dokter'))

@app.route('/kasir/riwayat', methods=['GET'])
@login_required
def kasir_riwayat():
    # Only kasir dapat akses
    if current_user.role != 'kasir':
        flash('Anda tidak memiliki akses', 'danger')
        return redirect(url_for('index'))
    
    search = request.args.get('search', '')
    dokter_id = request.args.get('dokter_id', '')
    
    query = PenggunaanKupon.query
    
    if dokter_id:
        query = query.filter_by(dokter_id=dokter_id)
    
    if search:
        query = query.filter(
            (PenggunaanKupon.user_email.ilike(f'%{search}%')) |
            (Dokter.nama_dokter.ilike(f'%{search}%'))
        ).join(Dokter)
    
    riwayat = query.order_by(PenggunaanKupon.tanggal_penggunaan.desc()).all()
    dokter_options = Dokter.query.filter_by(is_active=True).all()
    
    # Calculate stats
    total_kupon = sum([r.jumlah_kupon_digunakan for r in riwayat])
    
    return render_template('kasir_riwayat.html', 
                         riwayat=riwayat,
                         search=search,
                         dokter_id=dokter_id,
                         dokter_options=dokter_options,
                         total_kupon=total_kupon)

@app.route('/kasir/riwayat/export', methods=['GET'])
@login_required
def kasir_export_riwayat():
    # Only kasir dapat akses
    if current_user.role != 'kasir':
        flash('Anda tidak memiliki akses', 'danger')
        return redirect(url_for('index'))
    
    search = request.args.get('search', '')
    dokter_id = request.args.get('dokter_id', '')
    
    query = PenggunaanKupon.query
    
    if dokter_id:
        query = query.filter_by(dokter_id=dokter_id)
    
    if search:
        query = query.filter(
            (PenggunaanKupon.user_email.ilike(f'%{search}%')) |
            (Dokter.nama_dokter.ilike(f'%{search}%'))
        ).join(Dokter)
    
    riwayat = query.order_by(PenggunaanKupon.tanggal_penggunaan.desc()).all()
    
    # Create CSV
    output = StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow([
        'No', 'Nama Dokter', 'Spesialisasi', 'Jumlah Kupon', 'Tanggal Potong', 'Catatan'
    ])
    
    # Write data
    for idx, item in enumerate(riwayat, 1):
        writer.writerow([
            idx,
            item.dokter.nama_dokter,
            item.dokter.spesialisasi,
            item.jumlah_kupon_digunakan,
            item.tanggal_penggunaan.strftime('%d/%m/%Y %H:%M'),
            item.catatan or ''
        ])
    
    # Create bytes response
    output.seek(0)
    mem = BytesIO()
    mem.write(output.getvalue().encode('utf-8-sig'))
    mem.seek(0)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'laporan_kupon_kasir_{timestamp}.csv'
    
    return send_file(
        mem,
        mimetype='text/csv',
        as_attachment=True,
        download_name=filename
    )

@app.route('/admin/laporan-kupon', methods=['GET'])
@login_required
def admin_laporan_kupon():
    # Only admin dapat akses
    if current_user.role != 'admin':
        flash('Anda tidak memiliki akses', 'danger')
        return redirect(url_for('dashboard'))
    
    search = request.args.get('search', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    query = PenggunaanKupon.query
    
    if search:
        query = query.filter(
            (PenggunaanKupon.user_email.ilike(f'%{search}%')) |
            (Dokter.nama_dokter.ilike(f'%{search}%'))
        ).join(Dokter)
    
    # Date range filtering
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(PenggunaanKupon.tanggal_penggunaan >= date_from_obj)
        except:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            date_to_obj = date_to_obj.replace(hour=23, minute=59, second=59)
            query = query.filter(PenggunaanKupon.tanggal_penggunaan <= date_to_obj)
        except:
            pass
    
    riwayat = query.order_by(PenggunaanKupon.tanggal_penggunaan.desc()).all()
    
    # Calculate statistics
    total_pemakaian = sum([r.jumlah_kupon_digunakan for r in riwayat])
    
    # Group by doctor for summary
    dokter_summary = {}
    for item in riwayat:
        doc_name = item.dokter.nama_dokter
        if doc_name not in dokter_summary:
            dokter_summary[doc_name] = {
                'total': 0,
                'spesialisasi': item.dokter.spesialisasi,
                'jumlah_kupon_awal': item.dokter.kupon_awal or 0
            }
        dokter_summary[doc_name]['total'] += item.jumlah_kupon_digunakan
    
    return render_template('admin_laporan_kupon.html',
                         riwayat=riwayat,
                         dokter_summary=dokter_summary,
                         total_pemakaian=total_pemakaian,
                         search=search,
                         date_from=date_from,
                         date_to=date_to)

@app.route('/admin/laporan-kupon/export', methods=['GET'])
@login_required
def admin_export_laporan_kupon():
    # Only admin dapat akses
    if current_user.role != 'admin':
        flash('Anda tidak memiliki akses', 'danger')
        return redirect(url_for('dashboard'))
    
    search = request.args.get('search', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    query = PenggunaanKupon.query
    
    if search:
        query = query.filter(
            (PenggunaanKupon.user_email.ilike(f'%{search}%')) |
            (Dokter.nama_dokter.ilike(f'%{search}%'))
        ).join(Dokter)
    
    # Date range filtering
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(PenggunaanKupon.tanggal_penggunaan >= date_from_obj)
        except:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            date_to_obj = date_to_obj.replace(hour=23, minute=59, second=59)
            query = query.filter(PenggunaanKupon.tanggal_penggunaan <= date_to_obj)
        except:
            pass
    
    riwayat = query.order_by(PenggunaanKupon.tanggal_penggunaan.desc()).all()
    
    # Create CSV
    output = StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow([
        'No', 'Nama Dokter', 'Spesialisasi', 'Member/Nama', 'Jumlah Kupon', 'Tanggal Pemakaian', 'Catatan'
    ])
    
    # Write data
    for idx, item in enumerate(riwayat, 1):
        writer.writerow([
            idx,
            item.dokter.nama_dokter,
            item.dokter.spesialisasi,
            item.user_email,
            item.jumlah_kupon_digunakan,
            item.tanggal_penggunaan.strftime('%d/%m/%Y %H:%M'),
            item.catatan or ''
        ])
    
    # Create bytes response
    output.seek(0)
    mem = BytesIO()
    mem.write(output.getvalue().encode('utf-8-sig'))
    mem.seek(0)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'laporan_pemakaian_kupon_{timestamp}.csv'
    
    return send_file(
        mem,
        mimetype='text/csv',
        as_attachment=True,
        download_name=filename
    )

@app.route('/admin/users', methods=['GET', 'POST'])
@login_required
def admin_users():
    """Manage system users"""
    if current_user.role != 'admin':
        flash('Anda tidak memiliki akses', 'danger')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'add':
            email = request.form.get('email', '').strip()
            role = request.form.get('role', 'user')
            
            # Validasi email
            if not email or '@' not in email:
                flash('Email tidak valid', 'danger')
                return redirect(url_for('admin_users'))
            
            # Check if user exists
            if User.query.filter_by(email=email).first():
                flash(f'Email {email} sudah terdaftar', 'danger')
                return redirect(url_for('admin_users'))
            
            try:
                # Generate random password
                import secrets
                temp_password = secrets.token_urlsafe(8)
                
                # Create user
                user = User(email=email, role=role)
                user.set_password(temp_password)
                db.session.add(user)
                db.session.commit()
                
                # Send welcome email with credentials
                subject = f"Akun Baru - {role.upper()} | Sistem Koperasi"
                body = f"""
                <html>
                <body style="font-family: Arial, sans-serif;">
                    <h3 style="color: #1e5a96;">Selamat Datang di Sistem Koperasi!</h3>
                    <p>Akun Anda telah dibuat oleh administrator.</p>
                    
                    <h4 style="color: #1e5a96; margin-top: 20px;">Informasi Login:</h4>
                    <table style="border-collapse: collapse; width: 100%; margin-bottom: 20px;">
                        <tr>
                            <td style="padding: 8px; font-weight: bold; width: 30%;">Email:</td>
                            <td style="padding: 8px; background-color: #f0f0f0; font-family: monospace;">{email}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px; font-weight: bold;">Password Sementara:</td>
                            <td style="padding: 8px; background-color: #f0f0f0; font-family: monospace;">{temp_password}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px; font-weight: bold;">Role:</td>
                            <td style="padding: 8px;">{role.upper()}</td>
                        </tr>
                    </table>
                    
                    <p style="color: #d32f2f;"><strong>⚠️ PENTING:</strong></p>
                    <ul>
                        <li>Gunakan email dan password di atas untuk login pertama kali</li>
                        <li>Segera ubah password Anda setelah login</li>
                        <li>Jangan bagikan password ini ke orang lain</li>
                    </ul>
                    
                    <p style="margin-top: 20px;">
                        <a href="http://127.0.0.1:5000/login" style="display: inline-block; background-color: #1e5a96; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">Login ke Sistem</a>
                    </p>
                    
                    <hr style="margin: 20px 0; border: none; border-top: 1px solid #ddd;">
                    <p style="font-size: 12px; color: #999;">Email ini dikirim otomatis dari sistem Aplikasi Koperasi.</p>
                </body>
                </html>
                """
                
                try:
                    send_email(email, subject, body)
                    flash(f'✓ User {email} berhasil ditambahkan. Email login telah dikirim.', 'success')
                except Exception as e:
                    flash(f'User ditambahkan tapi gagal mengirim email: {str(e)}', 'warning')
                
            except Exception as e:
                db.session.rollback()
                flash(f'Gagal menambahkan user: {str(e)}', 'danger')
        
        elif action == 'delete':
            user_id = request.form.get('user_id')
            user = User.query.get(user_id)
            
            if not user:
                flash('User tidak ditemukan', 'danger')
            elif user.id == current_user.id:
                flash('Anda tidak bisa menghapus akun sendiri', 'danger')
            else:
                try:
                    email = user.email
                    db.session.delete(user)
                    db.session.commit()
                    flash(f'✓ User {email} berhasil dihapus', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Gagal menghapus user: {str(e)}', 'danger')
        
        elif action == 'reset_password':
            user_id = request.form.get('user_id')
            user = User.query.get(user_id)
            
            if not user:
                flash('User tidak ditemukan', 'danger')
            else:
                try:
                    import secrets
                    temp_password = secrets.token_urlsafe(8)
                    user.set_password(temp_password)
                    db.session.commit()
                    
                    # Send email with new password
                    subject = "Reset Password - Sistem Koperasi"
                    body = f"""
                    <html>
                    <body style="font-family: Arial, sans-serif;">
                        <h3 style="color: #1e5a96;">Password Anda Telah Direset</h3>
                        <p>Administrator telah mereset password Anda.</p>
                        
                        <table style="border-collapse: collapse; width: 100%; margin: 20px 0;">
                            <tr>
                                <td style="padding: 8px; font-weight: bold;">Password Baru:</td>
                                <td style="padding: 8px; background-color: #f0f0f0; font-family: monospace;">{temp_password}</td>
                            </tr>
                        </table>
                        
                        <p style="color: #d32f2f;"><strong>⚠️ Segera ubah password setelah login!</strong></p>
                        
                        <p style="margin-top: 20px;">
                            <a href="http://127.0.0.1:5000/login" style="display: inline-block; background-color: #1e5a96; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">Login</a>
                        </p>
                    </body>
                    </html>
                    """
                    
                    try:
                        send_email(user.email, subject, body)
                        flash(f'✓ Password {user.email} direset. Email telah dikirim.', 'success')
                    except Exception as e:
                        flash(f'Password direset tapi gagal mengirim email: {str(e)}', 'warning')
                
                except Exception as e:
                    db.session.rollback()
                    flash(f'Gagal reset password: {str(e)}', 'danger')
        
        return redirect(url_for('admin_users'))
    
    # Get all users
    users = User.query.all()
    role_options = ['user', 'HC', 'KaKop', 'admin', 'kasir']
    
    return render_template('admin_users.html', users=users, role_options=role_options, current_user_id=current_user.id)

@app.route('/uploads/<path:filename>')
@login_required
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

def init_db():
    with app.app_context():
        db.create_all()
        
        # Create default users if they don't exist
        if not User.query.filter_by(email='approval1@example.com').first():
            user1 = User(email='approval1@example.com', role='approval1')
            user1.set_password('password123')
            db.session.add(user1)
        
        if not User.query.filter_by(email='approval2@example.com').first():
            user2 = User(email='approval2@example.com', role='approval2')
            user2.set_password('password123')
            db.session.add(user2)
        
        if not User.query.filter_by(email='admin@example.com').first():
            admin = User(email='admin@example.com', role='admin')
            admin.set_password('password123')
            db.session.add(admin)
        
        if not User.query.filter_by(email='kasir@example.com').first():
            kasir = User(email='kasir@example.com', role='kasir')
            kasir.set_password('password123')
            db.session.add(kasir)
        
        db.session.commit()
        print("Database initialized with default users")

if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000)