from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

# Get the correct path
base_path = r"c:\Users\agung.daniel\Project PBO\aplikasi formulir pendaftaran"
output_file = os.path.join(base_path, "Contoh_Format_Upload_Dokter.xlsx")

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = 'Data Dokter'

# Set column widths
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 30

# Define styles
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True, size=11)
border_style = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Create header row
headers = ['Nama Dokter', 'Nomor Kontak', 'Jumlah Kupon', 'Keterangan']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_style

# Add sample data
sample_data = [
    ['dr. Agung Wijaya, SpB', '081806996558', 5, 'Spesialis Bedah'],
    ['dr. Siti Nurhaliza, SpPD', '082145789123', 3, 'Spesialis Penyakit Dalam'],
    ['dr. Budi Santoso, SpJP', '083456789012', 4, 'Spesialis Jantung'],
    ['dr. Ratna Dewi, SpA', '084567890123', 6, 'Spesialis Anak'],
]

for row_idx, row_data in enumerate(sample_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border_style
        if col_idx == 3:
            cell.alignment = Alignment(horizontal='center', vertical='center')

# Freeze header row
ws.freeze_panes = 'A2'

# Save file
wb.save(output_file)
print(f'✓ File berhasil dibuat: {output_file}')
print(f'✓ File size: {os.path.getsize(output_file)} bytes')
